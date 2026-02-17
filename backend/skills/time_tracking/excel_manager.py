"""
Master Excel timesheet management.

Manages monthly Excel files in ledger format (one row per time entry).
Files are named YYYY-MM-timesheet.xlsx and stored in the configured directory.
"""

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

COLUMNS = ["Date", "Project", "Project ID", "Start", "End", "Hours", "Description"]


def _file_path_for_date(entry_date: str, timesheet_dir: str) -> Path:
    """Get the Excel file path for a given date's month."""
    d = datetime.strptime(entry_date, "%Y-%m-%d")
    return Path(timesheet_dir) / f"{d.strftime('%Y-%m')}-timesheet.xlsx"


def _ensure_workbook(file_path: Path) -> Workbook:
    """Load existing workbook or create new one with headers."""
    if file_path.exists():
        return load_workbook(file_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.append(COLUMNS)
    return wb


def _entry_to_row(entry: dict) -> list:
    """Convert an entry dict to a spreadsheet row."""
    return [
        entry["date"],
        entry.get("project", ""),
        entry.get("project_id", ""),
        entry.get("start", ""),
        entry.get("end", ""),
        entry.get("hours", 0),
        entry.get("description", ""),
    ]


def log_entries(entries: list[dict], timesheet_dir: str) -> dict:
    """
    Append time entries to the master Excel file.

    Each entry dict should have: date, project, project_id, hours.
    Optional: start, end, description.

    Returns dict with success status and metadata.
    """
    dir_path = Path(timesheet_dir)
    dir_path.mkdir(parents=True, exist_ok=True)

    # Group entries by target file to avoid repeated open/save per entry
    by_file: dict[Path, list[dict]] = defaultdict(list)
    for entry in entries:
        file_path = _file_path_for_date(entry["date"], timesheet_dir)
        by_file[file_path].append(entry)

    for file_path, file_entries in by_file.items():
        wb = _ensure_workbook(file_path)
        ws = wb.active
        for entry in file_entries:
            ws.append(_entry_to_row(entry))
        wb.save(file_path)

    return {
        "success": True,
        "file": ", ".join(sorted(p.name for p in by_file)),
        "entries_written": len(entries),
    }


def _months_in_range(start_date: str, end_date: str) -> list[str]:
    """Return sorted YYYY-MM strings for every month spanned by the date range."""
    current = datetime.strptime(start_date, "%Y-%m-%d").date().replace(day=1)
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    months: list[str] = []
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    return months


def _parse_row_date(raw_value) -> str | None:
    """Parse a cell value into a YYYY-MM-DD string, or None if unparseable."""
    if raw_value is None:
        return None
    date_str = str(raw_value).split(" ")[0]  # handle datetime objects
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None
    return date_str


def _row_to_entry(row: tuple) -> dict:
    """Convert a spreadsheet row tuple to an entry dict."""
    return {
        "date": str(row[0]).split(" ")[0],
        "project": str(row[1] or ""),
        "project_id": str(row[2] or ""),
        "start": str(row[3] or ""),
        "end": str(row[4] or ""),
        "hours": float(row[5]) if row[5] else 0.0,
        "description": str(row[6] or ""),
    }


def read_entries(
    start_date: str,
    end_date: str,
    timesheet_dir: str,
    project_id: str | None = None,
) -> list[dict]:
    """
    Read time entries from master Excel for a date range.

    Spans multiple monthly files if the range crosses month boundaries.
    Optionally filters by project_id.
    """
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    dir_path = Path(timesheet_dir)
    results: list[dict] = []

    for month in _months_in_range(start_date, end_date):
        file_path = dir_path / f"{month}-timesheet.xlsx"
        if not file_path.exists():
            continue

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            date_str = _parse_row_date(row[0])
            if date_str is None:
                continue

            row_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            if row_date < start or row_date > end:
                continue

            entry = _row_to_entry(row)
            if project_id and entry["project_id"] != project_id:
                continue

            results.append(entry)

        wb.close()

    return results
